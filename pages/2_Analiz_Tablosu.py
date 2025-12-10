import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Sınıf Analizi", layout="wide")

st.title("📊 Sınıf Analizi ve Raporlama")

# 1. HAFIZA KONTROLÜ
if 'sinif_verileri' not in st.session_state or len(st.session_state.sinif_verileri) == 0:
    st.info("Henüz veri yok. Lütfen Ana Sayfa'dan kağıt okutun.")
    st.stop()

# 2. VERİYİ HAZIRLA
df = pd.DataFrame(st.session_state.sinif_verileri)
gosterilecek_df = df.drop(columns=["Detaylar"], errors="ignore")

# 3. İSTATİSTİKLER PANELI
col1, col2, col3, col4 = st.columns(4)
col1.metric("Öğrenci Sayısı", len(df))
if 'Toplam Puan' in df.columns:
    col2.metric("Ortalama", f"{df['Toplam Puan'].mean():.1f}")
    col3.metric("En Yüksek", df['Toplam Puan'].max())
    col4.metric("En Düşük", df['Toplam Puan'].min())

st.markdown("---")

# 4. WEB TABLOSU
st.subheader("📋 Detaylı Liste")
st.dataframe(gosterilecek_df, use_container_width=True)

# 5. PROFESYONEL EXCEL MOTORU (RENKLİ VE TASARIMLI)
def create_styled_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Veriyi yaz
        df.to_excel(writer, index=False, sheet_name='Sinav_Raporu')
        workbook = writer.book
        worksheet = writer.sheets['Sinav_Raporu']

        # --- TASARIM AYARLARI ---
        
        # 1. Başlık Stili (Koyu Mavi Zemin, Beyaz Kalın Yazı)
        header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Başlıkları Boya
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 2. Veri Hücreleri ve Renklendirme
        # Sütun isimlerini alalım ki hangi sütun "Puan" içeriyor bilelim
        columns = df.columns.tolist()
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                
                # Sütun başlığını bul
                col_name = columns[cell.col_idx - 1]
                
                # Eğer bu bir PUAN sütunuysa renklendirelim
                if "Soru" in col_name or "Toplam" in col_name:
                    try:
                        val = float(cell.value)
                        # Yeşil (Yüksek Puan - Örn: 80 üstü veya Tam puan gibi)
                        # Burada basitçe: 0 ise Kırmızı, Değilse normal bırakıyoruz
                        if val == 0:
                            cell.font = Font(color="FF0000", bold=True) # Kırmızı Yazı
                        elif val > 0:
                            cell.font = Font(color="006100", bold=True) # Koyu Yeşil Yazı
                    except:
                        pass

        # 3. Sütun Genişliklerini Otomatik Ayarla
        for i, column in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(i + 1)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output.getvalue()

# ==========================================
# İNDİRME BUTONU
# ==========================================
st.markdown("---")
col_d1, col_d2 = st.columns([2,1])

with col_d1:
    st.info("💡 İpucu: İndirilen Excel dosyasında, notlar web'deki gibi renklendirilmiş ve hizalanmış olacaktır.")

with col_d2:
    excel_data = create_styled_excel(gosterilecek_df)
    st.download_button(
        label="📥 Raporu Renkli Excel Olarak İndir",
        data=excel_data,
        file_name='Sinav_Analiz_Raporu.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type="primary",
        use_container_width=True
    )
